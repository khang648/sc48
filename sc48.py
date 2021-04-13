########################################################## Import Module ################################################################
from tkinter import *
from tkinter import messagebox
import time
from time import sleep, gmtime, strftime
from picamera import PiCamera
import cv2 
import numpy as np
from tkinter import filedialog as fd
from PIL import ImageTk, Image
import serial
from functools import partial
import math
from fractions import Fraction
from threading import *
import os
from tkinter import ttk
import awesometkinter as atk
import tkinter.font as font
import openpyxl
from openpyxl import Workbook, load_workbook
import subprocess
import shutil
# from gpiozero import LED
# tx = LED(14)
# rx = LED(15)
# rx.off()
# tx.off()
# sleep(1)

######################################################### Khởi tạo Main Window ##############################################################
root = Tk()
root.title(" ")
root.geometry('1024x600')
root.configure(background = "white")
root.attributes('-fullscreen', True)
root.resizable(False,False)
def disable_event():
    pass
root.protocol("WM_DELETE_WINDOW", disable_event)
s = ttk.Style()
s.theme_use('clam')

########################################################## Biến toàn cục ################################################################# 
covid19clicked = 0
tbclicked = 0
spotcheckclicked = 0
shrimpclicked = 0
viewresultclicked = 0 
temp_label = 0
name = "/"
entry_num = 0
wait = 0
pos_result = list(range(48))
path0 = "/"
path1 = "/"
path2 = "/"
path3 = "/"
path4 = "/"
path5 = "/"
foldername = ""
samples = 0
covid19_createclicked = 0
tb_createclicked = 0
spotcheck_createclicked = 0
shrimp_createclicked = 0
covid19dir_old = ""
tbdir_old = ""
spotcheckdir_old = ""
shrimpdir_old = ""
div = list(range(48))
start_point = (0,0)
end_point = (0,0)
calibrationclicked = 0
coefficient_value = list(range(48))

workbook = openpyxl.load_workbook('/home/pi/Desktop/sc48/Calibration_Value.xlsx')
sheet = workbook.active
for i in range(0,48):
    if(i<6):
        pos = str(chr(65+i+1)) + "2"
    if(i>=6 and i<12):
        pos = str(chr(65+i-5)) + "3"
    if(i>=12 and i<18):
        pos = str(chr(65+i-11)) + "4"
    if(i>=18 and i<24):
        pos = str(chr(65+i-17)) + "5"
    if(i>=24 and i<30):
        pos = str(chr(65+i-23)) + "6"
    if(i>=30 and i<36):
        pos = str(chr(65+i-29)) + "7"
    if(i>=36 and i<42):
        pos = str(chr(65+i-35)) + "8"
    if(i>=42):
        pos = str(chr(65+i-41)) + "9"
        
    coefficient_value[i] = sheet[pos].value
print('Coefficient Value: ',coefficient_value)

################################################################ Camera ####################################################################
def camera_capture(output):
    camera = PiCamera(framerate=Fraction(1,6), sensor_mode=3)
    camera.rotation = 180
    camera.iso = 800
    sleep(2)
    camera.shutter_speed = 6000000
    camera.exposure_mode = 'off'
    #sleep(20)
    camera.capture(output)
    camera.close()

########################################################### Sắp xếp contours ###############################################################
def sorting_y(contour):
    rect_y = cv2.boundingRect(contour)
    return rect_y[1]
def sorting_x(contour):
    rect_x = cv2.boundingRect(contour)
    return rect_x[0]
def sorting_xy(contour):
    rect_xy = cv2.boundingRect(contour)
    return math.sqrt(math.pow(rect_xy[0],2) + math.pow(rect_xy[1],2))

############################################################# Xử lý ảnh ###################################################################
def process_image(image_name, start_point=(282,79), end_point=(515,392)):
    global coefficient_value
    global calibrationclicked
    
    image = cv2.imread(image_name)
    blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
    #blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,10,10,7,21)
    gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)            
    thresh, binary_img = cv2.threshold(gray_img.copy(), 30, maxval=255, type=cv2.THRESH_BINARY) 
    contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print("Number of contours: " + str(len(contours)))

    contours.sort(key=lambda data:sorting_xy(data))

    contour_img = np.zeros_like(gray_img)
    contour_img = cv2.rectangle(contour_img, start_point, end_point, (255,255,255), -1)
    rect_w = end_point[0] - start_point[0]
    rect_h = end_point[1] - start_point[1]
    cell_w = round(rect_w/6)
    cell_h = round(rect_h/8)
    for i in range(1,6):
        contour_img = cv2.line(contour_img, (start_point[0]+i*cell_w,start_point[1]), (start_point[0]+i*cell_w,end_point[1]),(0,0,0), 2)
    for i in range(1,8):
        contour_img = cv2.line(contour_img, (start_point[0],start_point[1]+i*cell_h), (end_point[0],start_point[1]+i*cell_h),(0,0,0), 2)

    #gray1_img = cv2.cvtColor(contour_img, cv2.COLOR_BGR2GRAY)
    thresh1 , binary1_img = cv2.threshold(contour_img, 250, maxval=255, type=cv2.THRESH_BINARY)
    contours1, hierarchy1 = cv2.findContours(binary1_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)

    contours1.sort(key=lambda data:sorting_y(data))
    contours1_h1 = contours1[0:6]
    contours1_h2 = contours1[6:12]
    contours1_h3 = contours1[12:18]
    contours1_h4 = contours1[18:24]
    contours1_h5 = contours1[24:30]
    contours1_h6 = contours1[30:36]
    contours1_h7 = contours1[36:42]
    contours1_h8 = contours1[42:48]
    contours1_h1.sort(key=lambda data:sorting_x(data))
    contours1_h2.sort(key=lambda data:sorting_x(data))
    contours1_h3.sort(key=lambda data:sorting_x(data))
    contours1_h4.sort(key=lambda data:sorting_x(data))
    contours1_h5.sort(key=lambda data:sorting_x(data))
    contours1_h6.sort(key=lambda data:sorting_x(data))
    contours1_h7.sort(key=lambda data:sorting_x(data))
    contours1_h8.sort(key=lambda data:sorting_x(data))

    sorted_contours1 = contours1_h1 + contours1_h2 + contours1_h3 + contours1_h4 + contours1_h5 + contours1_h6 + contours1_h7 + contours1_h8
                 
    list_intensities = []
    sum_intensities = []
    result_list = list(range(48))
    area = list(range(48))
    
#     grayprocess_img = cv2.cvtColor(image.copy(), cv2.COLOR_BGR2GRAY)
#     for i in range(len(sorted_contours1)):
#         cimg = np.zeros_like(gray_img)
#         cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
#         pts = np.where(cimg == 255)
#         list_intensities.append(grayprocess_img[pts[0], pts[1]])
#         sum_intensities.append(sum(list_intensities[i]))
#         area[i]= cv2.contourArea(sorted_contours1[i])
#         result_list[i] = round((sum_intensities[i])*50/69791)

    tmp_list = list(range(48))
    #blur1_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,9,25) 
    #cv2.imwrite("mau1.jpg",blur1_img)
    blur1_img = cv2.GaussianBlur(image.copy(), (13,13), 0)
    cv2.imwrite("mau2.jpg",blur_img)
    hsv_img = cv2.cvtColor(blur1_img, cv2.COLOR_BGR2HSV)
    list_hsvvalue = []
    list_index = list(range(48))
    for i in range(len(sorted_contours1)):
        list_index[i] = []
        cimg = np.zeros_like(gray_img)
        cv2.drawContours(cimg, sorted_contours1, i, color = 255, thickness = -1)
        pts = np.where(cimg == 255)
        list_hsvvalue.append(hsv_img[pts[0], pts[1]])
        for j in range(len(list_hsvvalue[i])):
            list_index[i].append(list_hsvvalue[i][j][2])
        list_intensities.append(sum(list_index[i]))
        #area[i]= cv2.contourArea(sorted_contours1[i])
        result_list[i] = list_intensities[i]
        tmp_list[i] = list_intensities[i]/1000
        result_list[i] = round(tmp_list[i])
        
             
        #result_list[i] = round((list_intensities[i])*20/36880)
            
#lay hieu so voi gia tri khong mau           
#         if(calibrationclicked==0):
#             result_list[i] = result_list[i]-coefficient_value[i]
#             if(result_list[i]<0):
#                 result_list[i]=0

#     if(calibrationclicked==0):
#         for i in range(len(sorted_contours1)):
#             if(i==0 or i==6):
#                 result_list[i] = round(result_list[i]*0.75)
#             if(i==7 or i==8 or i==10 or i==12 or i==14 or i==47):
#                 result_list[i] = round(result_list[i]*0.789)
#             if(i==9 or i==13):
#                 result_list[i] = round(result_list[i]*0.833)
#             if(i==16 or i==18 or i==24 or i==36 or i==37 or i==40):
#                 result_list[i] = round(result_list[i]*0.882)
#             if(i==15 or i==19 or i==22 or i==25 or i==30 or i==38 or i==39):
#                 result_list[i] = round(result_list[i]*0.9375)
                

#nhan he so
        if(calibrationclicked==0):
            result_list[i] = round(result_list[i]*coefficient_value[i])
        else:
            result_list[i] = result_list[i]
            
#         for i in range(len(sorted_contours1)):
#             if(i==7 or i==38 or i==40):
#                 result_list[i] = round(result_list[i]*1.111)
#             if(i==8 or i==16 or i==30 or i==34):
#                 result_list[i] = round(result_list[i]*1.081)
#             if(i==9 or i==37):
#                 result_list[i] = round(result_list[i]*1.143)
#             if(i==10 or i==39):
#                 result_list[i] = round(result_list[i]*1.176)
    
#     for i in range(len(sorted_contours1)):
#         if(result_list[i]>99):
#             result_list[i]=99

    for i in range(len(sorted_contours1)):
        if ((i!=0) and ((i+1)%6==0)):
            print('%d' %(result_list[i]))
        else:
            print('%d' % (result_list[i]), end = ' | ')

    blurori_img = cv2.GaussianBlur(image.copy(), (25,25), 0)
    for i in range(len(sorted_contours1)):
        if(result_list[i]<=20):
            cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
        else:
            if(result_list[i] <= 25):
                cv2.drawContours(blurori_img, sorted_contours1, i, (255,255,0), thickness = 2)
            else:
                cv2.drawContours(blurori_img, sorted_contours1, i, (0,0,255), thickness = 2)
#     for i in range(len(contours)):
#         cv2.drawContours(blurori_img, contours, i, (255,255,255), thickness = 1)

    return (result_list, blurori_img)

########################################################## Giao diện chính ################################################################
def mainscreen():
    buttonFont = font.Font(family='Helvetica', size=10, weight='bold')
    global mainscreen_labelframe
    mainscreen_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    mainscreen_labelframe.place(x=0,y=0)
    sidebar_labelframe = LabelFrame(mainscreen_labelframe, bg='dodger blue', width=170, height=478)
    sidebar_labelframe.place(x=0,y=0)
    
    def home_click():
        subprocess.Popen(['killall','florence'])
        root.attributes('-fullscreen', True)
        home_canvas['bg'] = 'white'
        covid19_canvas['bg'] = 'dodger blue'
        tb_canvas['bg'] = 'dodger blue'
        calibration_canvas['bg'] = 'dodger blue'
        shrimp_canvas['bg'] = 'dodger blue'
        
        global covid19clicked
        covid19clicked = 0
        global tbclicked
        tbclicked = 0
        global shrimpclicked
        shrimpclicked = 0
        
        global covid19_createclicked, tb_createclicked, shrimp_createclicked
        covid19_createclicked = 0
        tb_createclicked = 0
        shrimp_createclicked = 0

        homemc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        homemc_labelframe.place(x=172,y=0)
            
        logo_img = Image.open('logo.png')
        logo_width, logo_height = logo_img.size
        scale_percent = 100
        width = int(logo_width * scale_percent / 100)
        height = int(logo_height * scale_percent / 100)
        display_img = logo_img.resize((width,height))
        image_select = ImageTk.PhotoImage(display_img)
        logo_label = Label(mainscreen_labelframe, bg='white',image=image_select)
        logo_label.image = image_select
        logo_label.place(x=250,y=25)
        
        def newprogram_click():
            logo_label.place_forget()
            try:
                newprogram_button.place_forget()
            except:
                pass
            global calibrationclicked
            calibrationclicked = 0
            global spotcheckclicked
            spotcheckclicked = 1
            enterframe_labelframe = LabelFrame(homemc_labelframe, bg='white', width=480, height=175)
            enterframe_labelframe.place(x=70,y=40)
            foldername_label = Label(homemc_labelframe, bg='white',text='Folder name:', fg='black', font=("Courier",14,'bold'))
            foldername_label.place(x=90,y=95)
            
            global spotcheck_createclicked
            global spotcheckdir_old
            directory = strftime("SPOTCHECK %y-%m-%d %H.%M.%S")
            if(spotcheck_createclicked == 0):
                directory_label = Label(homemc_labelframe, bg='grey94',text=directory)
                directory_label.place(x=240,y=60)
                spotcheckdir_old = directory
            else:
                directory_label = Label(homemc_labelframe, bg='grey94',text=spotcheckdir_old)
                directory_label.place(x=240,y=60)
                
            def enter_entry(event):
                root.attributes('-fullscreen', False)
                subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
            global foldername
            foldername_entry = Entry(homemc_labelframe,width=35)
            foldername_entry.insert(0,foldername)
            foldername_entry.bind("<Button-1>", enter_entry)
            foldername_entry.place(x=240,y=95)
            
            def create_click():
                global spotcheck_createclicked
                spotcheck_createclicked = 1
                global foldername
                foldername = foldername_entry.get()       
                name = strftime(foldername)
                global path0
                global spotcheckdir_old
                path0 = os.path.join("/home/pi/Desktop/spotcheck result", covid19dir_old +" "+ name)
                if(foldername_entry.get()==""):
                    msgbox = messagebox.showwarning(" ","Please enter the folder name !" )
                else:
                    if os.path.exists(path0):
                        subprocess.Popen(['killall','florence'])
                        root.attributes('-fullscreen', True)
                        msg = messagebox.askquestion("The folder already exixts", "Do you want to overwrite it?")
                        if(msg=='yes'):
                            shutil.rmtree(path0)
                            os.mkdir(path0)
                            global path1
                            path1 = os.path.join(path0,"Original image")
                            os.mkdir(path1)
                            global path2
                            path2 = os.path.join(path0,"Processed image")
                            os.mkdir(path2)
                            global path3
                            path3 = os.path.join(path0,"Result Table")
                            os.mkdir(path3)
                            global path4
                            path4 = os.path.join(path0,"Sample image")
                            os.mkdir(path4)
                            global path5
                            path5 = os.path.join(path0,"Temperature program")
                            os.mkdir(path5)
                            mainscreen_labelframe.place_forget()
                            settemp()
                    else:
                        subprocess.Popen(['killall','florence'])
                        root.attributes('-fullscreen', True)
                        os.mkdir(path0)
                        path1 = os.path.join(path0,"Original image")
                        os.mkdir(path1)
                        path2 = os.path.join(path0,"Processed image")
                        os.mkdir(path2)
                        path3 = os.path.join(path0,"Result Table")
                        os.mkdir(path3)
                        path4 = os.path.join(path0,"Sample image")
                        os.mkdir(path4)
                        path5 = os.path.join(path0,"Temperature program")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        settemp()
                            
            create_button = Button(homemc_labelframe, font=("Courier",12,'bold'), bg="lavender", text="CREATE", height=3, width=11, borderwidth=0, command=create_click)
            create_button.place(x=240,y=135)
        
        global spotcheck_createclicked
        if(spotcheck_createclicked==0):
            newprogram_button = Button(homemc_labelframe, font=("Courier",11,'bold'), bg="lavender", text="NEW PROGRAM", height=4, width=15, borderwidth=0, command=newprogram_click)
            newprogram_button.place(x=228,y=250)
            def shutdown_click():
                os.system("sudo shutdown -h now")
            def restart_click():
                os.system("sudo shutdown -r now")
            def exit_click():
                root.destroy()
            exit_button = Button(homemc_labelframe, fg='white', activebackground="dodger blue", font=('Courier','13','bold'), bg="blue4", text="EXIT", height=3, width=9, borderwidth=0, command=exit_click)
            exit_button.place(x=120,y=380)
            shutdown_button = Button(homemc_labelframe, fg='white', activebackground="red", font=('Courier','13','bold'), bg="red3", text="SHUTDOWN", height=3, width=9, borderwidth=0, command=shutdown_click)
            shutdown_button.place(x=250,y=380)
            restart_button = Button(homemc_labelframe, fg='white', activebackground="lawn green", font=('Courier','13','bold'), bg="green", text="RESTART", height=3, width=9, borderwidth=0, command=restart_click)
            restart_button.place(x=380,y=380)
        else:
            newprogram_click()
            
    def covid19_click():
        global calibrationclicked
        calibrationclicked = 0
        root.attributes('-fullscreen', True)
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'white'
        tb_canvas['bg'] = 'dodger blue'
        calibration_canvas['bg'] = 'dodger blue'
        shrimp_canvas['bg'] = 'dodger blue'
        
        global covid19clicked
        covid19clicked = 1
        global tbclicked
        tbclicked = 0
        global spotcheckclicked
        spotcheckclicked = 0
        global shrimpclicked
        shrimpclicked = 0
           
        global spotcheck_createclicked, tb_createclicked, shrimp_createclicked
        spotcheck_createclicked = 0
        tb_createclicked = 0
        shrimp_createclicked = 0
        
        covid19mc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        covid19mc_labelframe.place(x=172,y=0)
        
        enterframe_labelframe = LabelFrame(covid19mc_labelframe, bg='white', width=480, height=175)
        enterframe_labelframe.place(x=70,y=40)
        foldername_label = Label(covid19mc_labelframe, bg='white',text='Folder name:', fg='black', font=("Courier",14,'bold'))
        foldername_label.place(x=90,y=95)
        
        global covid19_createclicked
        global covid19dir_old
        directory = strftime("COVID19 %y-%m-%d %H.%M.%S")
        if(covid19_createclicked == 0):
            directory_label = Label(covid19mc_labelframe, bg='grey94',text=directory)
            directory_label.place(x=240,y=60)
            covid19dir_old = directory
        else:
            directory_label = Label(covid19mc_labelframe, bg='grey94',text=covid19dir_old)
            directory_label.place(x=240,y=60)
        
        def enter_entry(event):
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        global foldername
        foldername_entry = Entry(covid19mc_labelframe,width=35)
        foldername_entry.insert(0,foldername)
        foldername_entry.bind("<Button-1>", enter_entry)
        foldername_entry.place(x=240,y=95)
        
        def create_click():
            global covid19_createclicked
            covid19_createclicked = 1
            global foldername
            foldername = foldername_entry.get()       
            name = strftime(foldername)
            global path0
            global covid19dir_old
            path0 = os.path.join("/home/pi/Desktop/spotcheck result", covid19dir_old +" "+ name)
            if(foldername_entry.get()==""):
                msgbox = messagebox.showwarning(" ","Please enter the folder name !" )
            else:
                if os.path.exists(path0):
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    msg = messagebox.askquestion("The folder already exixts", "Do you want to overwrite it?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"Original image")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"Processed image")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"Result Table")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"Sample image")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Temperature program")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        settemp()
                else:
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"Original image")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"Processed image")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"Result Table")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"Sample image")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Temperature program")
                    os.mkdir(path5)
                    mainscreen_labelframe.place_forget()
                    settemp()
               
        create_button = Button(covid19mc_labelframe, font=("Courier",12,'bold'), bg="lavender", text="CREATE", height=3, width=11, borderwidth=0, command=create_click)
        create_button.place(x=240,y=135)
                   
    def tb_click():
        subprocess.Popen(['killall','florence'])
        global calibrationclicked
        calibrationclicked = 0
        root.attributes('-fullscreen', True)
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        tb_canvas['bg'] = 'white'
        calibration_canvas['bg'] = 'dodger blue'
        shrimp_canvas['bg'] = 'dodger blue'
        
        global covid19clicked
        covid19clicked = 0
        global tbclicked
        tbclicked = 1
        global spotcheckclicked
        spotcheckclicked = 0
        global shrimpclicked
        shrimpclicked = 0
           
        global spotcheck_createclicked, covid19_createclicked, shrimp_createclicked
        spotcheck_createclicked = 0
        covid19_createclicked = 0
        shrimp_createclicked = 0
    
        tbmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        tbmc_labelframe.place(x=172,y=0)
        
        enterframe_labelframe = LabelFrame(tbmc_labelframe, bg='white', width=480, height=175)
        enterframe_labelframe.place(x=70,y=40)
        foldername_label = Label(tbmc_labelframe, bg='white',text='Folder name:', fg='black', font=("Courier",14,'bold'))
        foldername_label.place(x=90,y=95)
        
        global tb_createclicked
        global tbdir_old
        directory = strftime("TB %y-%m-%d %H.%M.%S")
        if(tb_createclicked == 0):
            directory_label = Label(tbmc_labelframe, bg='grey94',text=directory)
            directory_label.place(x=240,y=60)
            tbdir_old = directory
        else:
            directory_label = Label(tbmc_labelframe, bg='grey94',text=tbdir_old)
            directory_label.place(x=240,y=60)
        
        def enter_entry(event):
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        global foldername
        foldername_entry = Entry(tbmc_labelframe,width=35)
        foldername_entry.insert(0,foldername)
        foldername_entry.bind("<Button-1>", enter_entry)
        foldername_entry.place(x=240,y=95)
        
        def create_click():
            global tb_createclicked
            tb_createclicked = 1
            global foldername
            foldername = foldername_entry.get()       
            name = strftime(foldername)
            global path0
            global tbdir_old
            path0 = os.path.join("/home/pi/Desktop/spotcheck result", tbdir_old +" "+ name)
            if(foldername_entry.get()==""):
                msgbox = messagebox.showwarning(" ","Please enter the folder name !" )
            else:
                if os.path.exists(path0):
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    msg = messagebox.askquestion("The folder already exixts", "Do you want to overwrite it?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"Original image")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"Processed image")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"Result Table")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"Sample image")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Temperature program")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        settemp()
                else:
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"Original image")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"Processed image")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"Result Table")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"Sample image")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Temperature program")
                    os.mkdir(path5)
                    mainscreen_labelframe.place_forget()
                    settemp()
                        
        create_button = Button(tbmc_labelframe, font=("Courier",12,'bold'), bg="lavender", text="CREATE", height=3, width=11, borderwidth=0, command=create_click)
        create_button.place(x=240,y=135)
        
    def calibration_click():
        subprocess.Popen(['killall','florence'])
        root.attributes('-fullscreen', True)
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        tb_canvas['bg'] = 'dodger blue'
        calibration_canvas['bg'] = 'white'
        shrimp_canvas['bg'] = 'dodger blue'
        
        global covid19clicked
        covid19clicked = 0
        global tbclicked
        tbclicked = 0
        global spotcheckclicked
        spotcheckclicked = 0
        global shrimpclicked
        shrimpclicked = 1
        
        calibrationmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        calibrationmc_labelframe.place(x=172,y=0)
        
        def calib_click():
            global calibrationclicked
            calibrationclicked = 1
            
            camera_capture('calib.jpg')
            
#             image = cv2.imread('calib.jpg')
#             #blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
#             blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,7,21)
#             gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)            
#             thresh, binary_img = cv2.threshold(gray_img.copy(), 40, maxval=255, type=cv2.THRESH_BINARY) 
#             contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
#             print("Number of contours: " + str(len(contours)))
# 
#             contours.sort(key=lambda data:sorting_xy(data))
# 
#             contour_img = np.zeros_like(gray_img)
#             bourect0 = cv2.boundingRect(contours[0])
#             bourect47 = cv2.boundingRect(contours[len(contours)-1])
#             global start_point
#             start_point = (bourect0[0]-6, bourect0[1]-6)
#             global end_point 
#             end_point = (bourect47[0]+bourect47[2]+6, bourect47[1]+bourect47[3]+6)
#             
            #calib_result, __ = process_image('calib.jpg')
            
            inten_result, __ = process_image('calib.jpg')
            calib_result = list(range(48))  
            for i in range(0,48):
                calib_result[i] = round(inten_result[20]/inten_result[i],3)
           
            for i in range(0,48):
                if((i!=0) and (i+1)%6==0):
                    print('%.3f' %(calib_result[i]))
                else:
                    print('%.3f' %(calib_result[i]), end= '|')
                    
            for i in range(0,48):
                if(i<6):
                    pos = str(chr(65+i+1)) + "2"
                if(i>=6 and i<12):
                    pos = str(chr(65+i-5)) + "3"
                if(i>=12 and i<18):
                    pos = str(chr(65+i-11)) + "4"
                if(i>=18 and i<24):
                    pos = str(chr(65+i-17)) + "5"
                if(i>=24 and i<30):
                    pos = str(chr(65+i-23)) + "6"
                if(i>=30 and i<36):
                    pos = str(chr(65+i-29)) + "7"
                if(i>=36 and i<42):
                    pos = str(chr(65+i-35)) + "8"
                if(i>=42):
                    pos = str(chr(65+i-41)) + "9"
                
                sheet[pos] = calib_result[i]
            workbook.save('Calibration_Value.xlsx')
            calibrationclicked  = 0
            msg = messagebox.showinfo('Calibration Done', 'Close the app and reboot!')
            root.destroy()
        calib_button = Button(calibrationmc_labelframe, font=("Courier",11,'bold'), bg="lavender", text="START CALIBRATION", height=4, width=15, borderwidth=0, command=calib_click)
        calib_button.place(x=230,y=320)
    
    def shrimp_click():
        global calibrationclicked
        calibrationclicked = 0
        root.attributes('-fullscreen', True)
        subprocess.Popen(['killall','florence'])
        root.attributes('-fullscreen', True)
        home_canvas['bg'] = 'dodger blue'
        covid19_canvas['bg'] = 'dodger blue'
        tb_canvas['bg'] = 'dodger blue'
        calibration_canvas['bg'] = 'dodger blue'
        shrimp_canvas['bg'] = 'white'
        
        global covid19clicked
        covid19clicked = 0
        global tbclicked
        tbclicked = 0
        global spotcheckclicked
        spotcheckclicked = 0
        global shrimpclicked
        shrimpclicked = 1
        
        global spotcheck_createclicked, tb_createclicked, covid19createclicked
        spotcheck_createclicked = 0
        tb_createclicked = 0
        covid19_createclicked = 0
        
        shrimpmc_labelframe = LabelFrame(mainscreen_labelframe, bg='white', width=624, height=478)
        shrimpmc_labelframe.place(x=172,y=0)
        
        enterframe_labelframe = LabelFrame(shrimpmc_labelframe, bg='white', width=480, height=175)
        enterframe_labelframe.place(x=70,y=40)
        foldername_label = Label(shrimpmc_labelframe, bg='white',text='Folder name:', fg='black', font=("Courier",14,'bold'))
        foldername_label.place(x=90,y=95)
        
        global shrimp_createclicked
        global shrimpdir_old
        directory = strftime("SHRIMP %y-%m-%d %H.%M.%S")
        if(shrimp_createclicked == 0):
            directory_label = Label(shrimpmc_labelframe, bg='grey94',text=directory)
            directory_label.place(x=240,y=60)
            shrimpdir_old = directory
        else:
            directory_label = Label(shrimpmc_labelframe, bg='grey94',text=shrimpdir_old)
            directory_label.place(x=240,y=60)
        
        def enter_entry(event):
            root.attributes('-fullscreen', False)
            subprocess.Popen('florence',stdout=subprocess.PIPE, shell=True)
        global foldername
        foldername_entry = Entry(shrimpmc_labelframe,width=35)
        foldername_entry.insert(0,foldername)
        foldername_entry.bind("<Button-1>", enter_entry)
        foldername_entry.place(x=240,y=95)
        
        def create_click():
            global shrimp_createclicked
            shrimp_createclicked = 1
            global foldername
            foldername = foldername_entry.get()       
            name = strftime(foldername)
            global path0
            global shrimpdir_old
            path0 = os.path.join("/home/pi/Desktop/spotcheck result", shrimpdir_old +" "+ name)
            if(foldername_entry.get()==""):
                msgbox = messagebox.showwarning(" ","Please enter the folder name !" )
            else:
                if os.path.exists(path0):
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    msg = messagebox.askquestion("The folder already exixts", "Do you want to overwrite it?")
                    if(msg=='yes'):
                        shutil.rmtree(path0)
                        os.mkdir(path0)
                        global path1
                        path1 = os.path.join(path0,"Original image")
                        os.mkdir(path1)
                        global path2
                        path2 = os.path.join(path0,"Processed image")
                        os.mkdir(path2)
                        global path3
                        path3 = os.path.join(path0,"Result Table")
                        os.mkdir(path3)
                        global path4
                        path4 = os.path.join(path0,"Sample image")
                        os.mkdir(path4)
                        global path5
                        path5 = os.path.join(path0,"Temperature program")
                        os.mkdir(path5)
                        mainscreen_labelframe.place_forget()
                        settemp()
                else:
                    subprocess.Popen(['killall','florence'])
                    root.attributes('-fullscreen', True)
                    os.mkdir(path0)
                    path1 = os.path.join(path0,"Original image")
                    os.mkdir(path1)
                    path2 = os.path.join(path0,"Processed image")
                    os.mkdir(path2)
                    path3 = os.path.join(path0,"Result Table")
                    os.mkdir(path3)
                    path4 = os.path.join(path0,"Sample image")
                    os.mkdir(path4)
                    path5 = os.path.join(path0,"Temperature program")
                    os.mkdir(path5)
                    mainscreen_labelframe.place_forget()
                    settemp()
               
        create_button = Button(shrimpmc_labelframe, font=("Courier",12,'bold'), bg="lavender", text="CREATE", height=3, width=11, borderwidth=0, command=create_click)
        create_button.place(x=240,y=135)
               
    home_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="HOME", fg='white', font=buttonFont, borderwidth=0, height=4, width=20,command=home_click)
    home_button.place(x=1,y=43)
    covid19_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="COVID 19", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=covid19_click)
    covid19_button.place(x=1,y=123)
    tb_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="TB", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=tb_click)
    tb_button.place(x=1,y=203)
    calibration_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="CALIBRATION", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=calibration_click)
    calibration_button.place(x=1,y=363)
    shrimp_button = Button(mainscreen_labelframe, bg="dodger blue", activebackground="dodger blue", text="SHRIMP", fg='white', font=buttonFont, borderwidth=0, height=4, width=20, command=shrimp_click)
    shrimp_button.place(x=1,y=283)
    
    home_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    home_canvas.place(x=1,y=45)
    covid19_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    covid19_canvas.place(x=1,y=125)
    tb_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    tb_canvas.place(x=1,y=205)
    calibration_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    calibration_canvas.place(x=1,y=365)
    shrimp_canvas = Canvas(mainscreen_labelframe, bg="dodger blue", bd=0, highlightthickness=0, height=72, width=13)
    shrimp_canvas.place(x=1,y=285)

    global covid19clicked
    global tbclicked
    if(covid19clicked==1):
        covid19_click()
    elif(tbclicked==1):
        tb_click()
    elif(shrimpclicked==1):
        shrimp_click()
    else:
        home_click() 
    
###################################################### Giao diện set nhiệt độ ############################################################ 
def settemp():
    if(covid19clicked==1):
        fr = open("covid19saved.txt","r")
    if(tbclicked==1):
        fr = open("tbsaved.txt","r")
    if(spotcheckclicked==1):
        fr = open("scsaved.txt","r")
    if(shrimpclicked==1):
        fr = open("shrimpsaved.txt","r")
    t1 = fr.readline()[3:5]
    t2 = fr.readline()[3:5]
    t3 = fr.readline()[3:5]
    thr1 = fr.readline()[5:7]
    thr2 = fr.readline()[5:7]
    thr3l = fr.readline()[6:8]
    thr3h = fr.readline()[6:8]
    global samples
    samples=0
    settemp_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    settemp_labelframe.place(x=0,y=0)
    settemptop_labelframe = LabelFrame(settemp_labelframe, bg='white', width=798, height=350)
    settemptop_labelframe.place(x=0,y=52)
    keypad_labelframe = LabelFrame(settemptop_labelframe, bg='white', width=285, height=323)
    keypad_labelframe.place(x=501,y=11)
    title_labelframe = LabelFrame(settemp_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    settemp_label = Label(settemp_labelframe, bg='dodger blue', fg='black', text='SET PARAMETER', font=("Courier",17,'bold'), width=20, height=1 )
    settemp_label.place(x=270,y=12)
    
    def numpad_click(btn):
        text = "%s" % btn
        if (text!="Delete" and text!="Default"):
            if(entry_num==1):
                t1_entry.insert(END, text)
            if(entry_num==2):
                t2_entry.insert(END, text)
            if(entry_num==3):
                t3_entry.insert(END, text)
            if(entry_num==4):
                thr1_entry.insert(END, text)
            if(entry_num==5):
                thr2_entry.insert(END, text)
            if(entry_num==6):
                thr3l_entry.insert(END, text)
            if(entry_num==7):
                thr3h_entry.insert(END, text)
        if text == 'Delete':
            if(entry_num==1):
                t1_entry.delete(0, END)
            if(entry_num==2):
                t2_entry.delete(0, END)
            if(entry_num==3):
                t3_entry.delete(0, END)
            if(entry_num==4):
                thr1_entry.delete(0, END)
            if(entry_num==5):
                thr2_entry.delete(0, END)
            if(entry_num==6):
                thr3l_entry.delete(0, END)
            if(entry_num==7):
                thr3h_entry.delete(0, END)
        if text == 'Default':
            if(entry_num==1):
                t1_entry.delete(0, END)
                t1_entry.insert(END, t1)
            if(entry_num==2):
                t2_entry.delete(0, END)
                t2_entry.insert(END, t2)
            if(entry_num==3):
                t3_entry.delete(0, END)
                t3_entry.insert(END, t3)
            if(entry_num==4):
                thr1_entry.delete(0, END)
                thr1_entry.insert(END, 25)
            if(entry_num==5):
                thr2_entry.delete(0, END)
                thr2_entry.insert(END, 25)
            if(entry_num==6):
                thr3l_entry.delete(0, END)
                thr3l_entry.insert(END, 25)
            if(entry_num==7):
                thr3h_entry.delete(0, END)
                thr3h_entry.insert(END, 25)
                
    def numpad():
        global numpad_labelframe
        numpad_labelframe = LabelFrame(keypad_labelframe, bg="white", width=385, height=395)
        numpad_labelframe.place(x=2,y=1)
        button_list = ['7',     '8',      '9',
                       '4',     '5',      '6',
                       '1',     '2',      '3',
                       '0',     'Delete', 'Default']
        r = 1
        c = 0
        n = 0
        btn = list(range(len(button_list)))
        for label in button_list:
            cmd = partial(numpad_click, label)
            btn[n] = Button(numpad_labelframe, text=label, font=font.Font(family='Helvetica', size=10, weight='bold'), width=9, height=4, command=cmd)
            btn[n].grid(row=r, column=c, padx=0, pady=0)
            n += 1
            c += 1
            if (c == 3):
                c = 0
                r += 1
    
    temp_labelframe = LabelFrame(settemptop_labelframe, text='TEMPERATURE', bg='white', width=490, height=180)
    temp_labelframe.place(x=3,y=2)
    thres_labelframe = LabelFrame(settemptop_labelframe, text='THRESHOLD', bg='white', width=490, height=149)
    thres_labelframe.place(x=3,y=185)
    
    cir_img = Image.open('cir.png')
    cir_width, cir_height = cir_img.size
    scale_percent = 14
    width = int(cir_width * scale_percent / 100)
    height = int(cir_height * scale_percent / 100)
    display_img = cir_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    t1cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t1cir_label.image = image_select
    t1cir_label.place(x=5,y=5)
    t2cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t2cir_label.image = image_select
    t2cir_label.place(x=170,y=5)
    t3cir_label = Label(temp_labelframe, bg='white', image=image_select)
    t3cir_label.image = image_select
    t3cir_label.place(x=335,y=5)
#     graycir_img = Image.open('graycir.png')
#     graycir_width, cir_height = cir_img.size
#     scale_percent = 16
#     width = int(cir_width * scale_percent / 100)
#     height = int(cir_height * scale_percent / 100)
#     display_img = graycir_img.resize((width,height))
#     image_select = ImageTk.PhotoImage(display_img)
#     t4cir_label = Label(settemptop_labelframe, bg='white', image=image_select)
#     t4cir_label.image = image_select
#     t4cir_label.place(x=275,y=175)
        
    def entryt1_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 1
        numpad()
    def entryt2_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 2
        numpad()
    def entryt3_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 3
        numpad()
    def entrythr1_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 4
        numpad()
    def entrythr2_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 5
        numpad()
    def entrythr3l_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 6
        numpad()
    def entrythr3h_click(event):
        global numpad_labelframe
        global entry_num
        entry_num = 7
        numpad()
        
    t1_label = Label(temp_labelframe, bg='white', text='T1', fg='black', font=("Courier",20,"bold"))
    t1_label.place(x=15, y=14)
    t1oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t1oc_label.place(x=107, y=55)
    t1_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t1_entry.place(x=47,y=50)
    t1_entry.bind('<Button-1>', entryt1_click)
    t1_entry.insert(0,t1)
    
    t2_label = Label(temp_labelframe, bg='white', text='T2', fg='black', font=("Courier",20,"bold"))
    t2_label.place(x=180, y=14)
    t2oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t2oc_label.place(x=272, y=55)
    t2_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t2_entry.place(x=212,y=50)
    t2_entry.bind('<Button-1>', entryt2_click)
    t2_entry.insert(0,t2)
    
    t3_label = Label(temp_labelframe, bg='white', text='T3', fg='black', font=("Courier",20,"bold"))
    t3_label.place(x=345, y=14)
    t3oc_label = Label(temp_labelframe, bg='white', text=chr(176)+'C', fg='red', font=("Courier", 11,"bold"))
    t3oc_label.place(x=437, y=55)
    t3_entry = Entry(temp_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",36,"bold"))
    t3_entry.place(x=377,y=50)
    t3_entry.bind('<Button-1>', entryt3_click)
    t3_entry.insert(0,t3)
    
    thr1_label = Label(thres_labelframe, bg='white', text='T1: ', fg='black', font=("Courier",24,"bold"))
    thr1_label.place(x=14, y=39)
    thr1_entry = Entry(thres_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
    thr1_entry.place(x=71,y=39)
    thr1_entry.bind('<Button-1>', entrythr1_click)
    thr1_entry.insert(0,thr1)

    thr2_label = Label(thres_labelframe, bg='white', text='T2: ', fg='black', font=("Courier",24,"bold"))
    thr2_label.place(x=165, y=39)
    thr2_entry = Entry(thres_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
    thr2_entry.place(x=227,y=39)
    thr2_entry.bind('<Button-1>', entrythr2_click)
    thr2_entry.insert(0,thr2)

    thr3l_label = Label(thres_labelframe, bg='white', text='T3-L: ', fg='black', font=("Courier",24,"bold"))
    thr3l_label.place(x=320, y=7)
    thr3l_entry = Entry(thres_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
    thr3l_entry.place(x=420,y=7)
    thr3l_entry.bind('<Button-1>', entrythr3l_click)
    thr3l_entry.insert(0,thr3l)

    thr3h_label = Label(thres_labelframe, bg='white', text='T3-H: ', fg='black', font=("Courier",24,"bold"))
    thr3h_label.place(x=320, y=71)
    thr3h_entry = Entry(thres_labelframe, width=2, justify='center', bg='white', borderwidth=0, fg ='grey32', font=("Courier",25,"bold"))
    thr3h_entry.place(x=420,y=71)
    thr3h_entry.bind('<Button-1>', entrythr3h_click)
    thr3h_entry.insert(0,thr3h)

#     t4_label = Label(settemptop_labelframe, bg = 'white', text='T4', fg='grey67', font=("Courier",20,"bold"))
#     t4_label.place(x=286, y=185)
    
    def back_click():
        settemp_labelframe.place_forget()
        mainscreen()
    def thread():
        th1 = Thread(target = next_click)
        th1.start()
    def next_click():
        settemp_labelframe.place_forget()
        global t1_set, t2_set, t3_set, thr1_set, thr2_set, thr3l_set, thr3h_set
        t1_set = t1_entry.get()[0:2]
        t2_set = t2_entry.get()[0:2]
        t3_set = t3_entry.get()[0:2]
        thr1_set = thr1_entry.get()[0:2]
        thr2_set = thr2_entry.get()[0:2]
        thr3l_set = thr3l_entry.get()[0:2]
        thr3h_set = thr3h_entry.get()[0:2]
        
        global path5
        if os.path.exists(path5+"/Temperature_program.txt"):
            fc= open(path5+"/Temperature_program.txt","w")
            fc.truncate(0)
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        else: 
            fc= open(path5+"/Temperature_program.txt","w+")
            fc.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fc.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fc.writelines("T3="+t3_entry.get()[0:2]+"\n")
        scanposition()
    def save_click():
        msg = messagebox.askquestion("Save Temperature", "Do you want to save?")
        if(msg=='yes'):
            if(covid19clicked==1):
                fw = open("covid19saved.txt","w")
            if(tbclicked==1):
                fw = open("tbsaved.txt","w")
            if(spotcheckclicked==1):
                fw = open("scsaved.txt","w")
            if(shrimpclicked==1):
                fw = open("shrimpsaved.txt","w")
            fw.truncate(0)
            fw.writelines("T1="+t1_entry.get()[0:2]+"\n")
            fw.writelines("T2="+t2_entry.get()[0:2]+"\n")
            fw.writelines("T3="+t3_entry.get()[0:2]+"\n")
            fw.writelines("THR1="+thr1_entry.get()[0:2]+"\n")
            fw.writelines("THR2="+thr2_entry.get()[0:2]+"\n")
            fw.writelines("THR3L="+thr3l_entry.get()[0:2]+"\n")
            fw.writelines("THR3H="+thr3h_entry.get()[0:2]+"\n")
        
    back_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Back" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)
    next_button = Button(settemp_labelframe, font=('Courier','12','bold'), bg="lavender", text="Next", height=3, width=11, borderwidth=0, command=thread)
    next_button.place(x=647,y=406)
    save_button = Button(settemp_labelframe, activebackground="gold", font=('Courier','12','bold'), bg="yellow", text="Save", height=3, width=11, borderwidth=0,command=save_click)
    save_button.place(x=332,y=406)

##################################################### Giao diện định vị mẫu ############################################################   
def scanposition():
    global path0
    global path1
    global path2
    global path3
    global path4
    global path5
    
    global ser
    ser.flushInput()
    ser.flushOutput()
    global scanpostion_labelframe
    scanposition_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    scanposition_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(scanposition_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    scanposition_label = Label(scanposition_labelframe, bg='dodger blue', text='SAMPLES POSITION', font=("Courier",17,'bold'), width=20, height=1 )
    scanposition_label.place(x=257,y=12)

    scan_img = Image.open('scan.png')
    scan_width, scan_height = scan_img.size
    scale_percent = 100
    width = int(scan_width * scale_percent / 100)
    height = int(scan_height * scale_percent / 100)
    display_img = scan_img.resize((width,height))
    image_select = ImageTk.PhotoImage(display_img)
    scan_label = Label(scanposition_labelframe, bg='white',image=image_select)
    scan_label.image = image_select
    scan_label.place(x=270,y=80)
    
    s = ttk.Style()
    s.theme_use('clam')
    s.configure("green.Horizontal.TProgressbar", foreground='green', background='green')
    scanposition_progressbar = ttk.Progressbar(root, orient = HORIZONTAL, style="green.Horizontal.TProgressbar", length = 200, mode = 'determinate')
    scanposition_progressbar.place(x=299,y=408)

    def back_click():
        try:
            camera.close()
        except Exception:
            pass
        scanposition_labelframe.place_forget()
        settemp()
    back_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Back" , height=3, width=11, borderwidth=0, command=back_click)
    back_button.place(x=14,y=406)

    send_data = 'P'
    ser.write(send_data.encode())

    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)
        scanposition_progressbar['value'] = 5
        root.update_idletasks()
        if(receive_data=='C'):
            global wait
            wait = 1
            scanposition_progressbar['value'] = 20
            root.update_idletasks()
            
    while(wait!=1):
        scanposition_progressbar['value'] = 2
        root.update_idletasks()
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)
            scanposition_progressbar['value'] = 10
            root.update_idletasks()
            if(receive_data=='C'):
                scanposition_progressbar['value'] = 20
                root.update_idletasks()
                wait = 1
                break;
    while(wait==1):
        process_label = Label(scanposition_labelframe, text='Processing...', bg='white', font=("Courier",13))
        process_label.place(x=333,y=440)
        try:
            camera_capture(path4 + "/Sample_original.jpg")
        except Exception as e :
            error = messagebox.askquestion("Error: "+ str(e), "Do you want to restart the divice?", icon = "error")
            if(error=='yes'):
                os.system("sudo shutdown -r now")
        
        #image = cv2.imread("test.jpg")
        image = cv2.imread(path4 + "/Sample_original.jpg")
        #blur_img = cv2.GaussianBlur(image.copy(), (35,35), 0)
        blur_img = cv2.fastNlMeansDenoisingColored(image.copy(),None,15,15,7,21)
        gray_img = cv2.cvtColor(blur_img, cv2.COLOR_BGR2GRAY)            
        thresh, binary_img = cv2.threshold(gray_img.copy(), 40, maxval=255, type=cv2.THRESH_BINARY) 
        contours, hierarchy = cv2.findContours(binary_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        print("Number of contours: " + str(len(contours)))

        contours.sort(key=lambda data:sorting_xy(data))

        contour_img = np.zeros_like(gray_img)
        bourect0 = cv2.boundingRect(contours[0])
        bourect47 = cv2.boundingRect(contours[len(contours)-1])
        global start_point
        start_point = (bourect0[0]-2, bourect0[1]-2)
        global end_point 
        end_point = (bourect47[0]+bourect47[2]+2, bourect47[1]+bourect47[3]+2)
        print('Start point:', start_point)
        print('End point:', end_point)
        scanposition_progressbar['value'] = 35
        root.update_idletasks()

        global pos_result
        #pos_result, pos_image = process_image("test.jpg", start_point, end_point)
        #pos_result, pos_image = process_image(path4 + "/Sample_original.jpg", start_point, end_point)
        pos_result, pos_image = process_image(path4 + "/Sample_original.jpg")
        scanposition_progressbar['value'] = 60
        root.update_idletasks()
        sleep(1)

        output = path4 + "/Sample_processed.jpg"
        cv2.imwrite(output, pos_image)
        scanposition_progressbar['value'] = 90
        root.update_idletasks()
        sleep(1)

        scanresult_labelframe = LabelFrame(scanposition_labelframe, bg='ghost white', width=528,height = 307)
        scanresult_labelframe.place(x=248,y=60)
        
        label = list(range(48))
        def result_table(range_a, range_b, row_value):
            global samples
            j=-1
            for i in range(range_a, range_b):
                j+=1
                if(i<6):
                    t='A'+ str(i+1)
                if(i>=6 and i<12):
                    t='B'+ str(i-5)
                if(i>=12 and i<18):
                    t='C'+ str(i-11)
                if(i>=18 and i<24):
                    t='D'+ str(i-17)
                if(i>=24 and i<30):
                    t='E'+ str(i-23)
                if(i>=30 and i<36):
                    t='F'+ str(i-29)
                if(i>=36 and i<42):
                    t='G'+ str(i-35)
                if(i>=42):
                    t='H'+ str(i-41)
                if(pos_result[i]< 15):
                    label[i] = Label(scanresult_labelframe, bg='gainsboro', text=t, width=5, height=2)
                    label[i].grid(row=row_value,column=j,padx=3,pady=3)
                else:
                    label[i] = Label(scanresult_labelframe, bg='OliveDrab1', text=t, width=5, height=2)
                    label[i].grid(row=row_value,column=j,padx=3,pady=3)
                    samples += 1
        scanposition_progressbar['value'] = 100
        root.update_idletasks()
        
        result_table(0,6,0)
        result_table(6,12,1)
        result_table(12,18,2)
        result_table(18,24,3)
        result_table(24,30,4)
        result_table(30,36,5)
        result_table(36,42,6)
        result_table(42,48,7)
        global samples
        samplenum_label = Label(scanposition_labelframe, text='Number of Samples: ' + str(samples), fg='dodger blue', bg='white', font=("Courier",13))
        samplenum_label.place(x=293,y=432)
        scan_label.place_forget()
        scanposition_progressbar.place_forget()
        process_label.place_forget()
        wait = 0
        samples = 0
        def thread():
            th1 = Thread(target = next_click)
            th1.start()
        def next_click():
            global createclicked
            createclicked = 0
            scanposition_labelframe.place_forget()   
            analysis()
        next_button = Button(scanposition_labelframe, font=("Courier",12,'bold'), bg="lavender", text="Next", height=3, width=11, borderwidth=0,command=thread)
        next_button.place(x=647,y=406)

################################################### Giao diện phân tích mẫu ###########################################################
def analysis():
    global ser
    ser.flushInput()
    ser.flushOutput()
    global analysis_labelframe
    analysis_labelframe = LabelFrame(root, bg='white', width=800, height=600)
    analysis_labelframe.place(x=0,y=0)
    title_labelframe = LabelFrame(analysis_labelframe, bg='dodger blue', width=798, height=50)
    title_labelframe.place(x=0,y=0)
    analysis_label = Label(analysis_labelframe, bg='dodger blue', text='SAMPLES ANALYSIS', font=("Courier",17,'bold'), width=20, height=1 )
    analysis_label.place(x=261,y=12)
    t_labelframe = LabelFrame(analysis_labelframe, bg='white', width=798, height=298)
    t_labelframe.place(x=0,y=70)

    t1_labelframe = LabelFrame(t_labelframe, bg='white',text="T1:"+t1_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t1_labelframe.place(x=0,y=2)
    t2_labelframe = LabelFrame(t_labelframe, bg='white',text="T2:"+t2_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t2_labelframe.place(x=199,y=2)
    t3_labelframe = LabelFrame(t_labelframe, bg='white',text="T3:"+t3_set+chr(176)+'C' , font=("Courier",13,'bold'), width=197, height=290)
    t3_labelframe.place(x=398,y=2)
    t4_labelframe = LabelFrame(t_labelframe, bg='white smoke',text="T4", width=197, height=290)
    t4_labelframe.place(x=597,y=2)
    t1wait_label = Label(t1_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t1wait_label.place(x=46,y=110)
    t2wait_label = Label(t2_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t2wait_label.place(x=46,y=110)
    t3wait_label = Label(t3_labelframe, text='...', fg='grey36', bg='white', font=("Courier",40,'bold'))
    t3wait_label.place(x=46,y=110)
    temp_label = Label(analysis_labelframe, bg='white', fg='grey36', font=("Courier",20,'bold'))
    temp_label.place(x=65,y=389)

    def stop_click():
        global ser
        msgbox = messagebox.askquestion('Stop the process','Do you want to stop the analysis ?', icon = 'question')
        if(msgbox=='yes'):
            send_data ='S'
            ser.write(send_data.encode())
            analysis_labelframe.place_forget()
            mainscreen()       
    
    def pause_click():
        global ser
        if(pause_button['text']=='Pause'):
            send_data ='P'
            ser.write(send_data.encode())
            pause_button['text']= 'Continue'
        else:
            send_data ='R'
            ser.write(send_data.encode())
            pause_button['text']= 'Pause'
    pause_button = Button(analysis_labelframe, bg="lavender", font=("Courier",12,'bold'), text="Pause" , height=3, width=10, borderwidth=0, command=pause_click)
    pause_button.place(x=450,y=390)
    stop_button = Button(analysis_labelframe, bg="lavender", font=("Courier",12,'bold'), text="STOP", height=3, width=10, borderwidth=0, command=stop_click)
    stop_button.place(x=590,y=390)
    root.update()
    
    send_data = "t"+ t1_set + "," + t2_set + "," + t3_set + "z"
    ser.write(send_data.encode())
    print("Data send: ", send_data)
    t0 = time.time()
    sleep(2)
  
    global wait
    if(ser.in_waiting>0):
        receive_data = ser.readline().decode('utf-8').rstrip()
        print("Data received:", receive_data)        
        if(receive_data=='Y'):
            autoprocess_label = Label(analysis_labelframe, bg='white', text="Program is processing...", fg='blue', font=("Courier",12,'bold'))
            autoprocess_label.place(x=65,y=438)
            wait = 1
    while(wait!=1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8').rstrip()
            print("Data received:", receive_data)        
            if(receive_data=='Y'):
                autoprocess_label = Label(analysis_labelframe, bg='white', text="Program is processing...", fg='blue', font=("Courier",12,'bold'))
                autoprocess_label.place(x=65,y=438)
                wait = 1
                break

    while(wait==1):
        if(ser.in_waiting>0):
            receive_data = ser.readline().decode('utf-8',errors='ignore').rstrip()
            #print("Data received:", receive_data)
            if(receive_data!='C1' and receive_data!='C2' and receive_data!='C3'):
                print("Data received:", receive_data)
                temp_label['text'] = 'Temperature: '+ receive_data + chr(176)+'C'
                root.update()
                
            if(receive_data=='C1'):
                print("Data received:", receive_data)
                t1wait_label.place_forget()
                t1_labelframe['bg'] = atk.DEFAULT_COLOR
                t1_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t1_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t1_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Processing!', font=("Courier",9,'bold'))
                tprocess_label.place(x=59,y=112)

                global path1
                camera_capture(path1 + "/T1.jpg")
                  
                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                
                global start_point
                global end_point 
                t1_result, t1_image= process_image(path1 + "/T1.jpg", start_point, end_point)
                #t1_result, t1_image= process_image(path1 + "/T1.jpg")

                global path2
                output = path2 + "/T1.jpg"
                cv2.imwrite(output, t1_image)

                t1_analysis = Image.open(output)
                t1_crop = t1_analysis.crop((start_point[0]-7, start_point[1]-7, end_point[0]+7, end_point[1]+7))   
                crop_width, crop_height = t1_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t1_crop.resize((width,height))
                t1_display = ImageTk.PhotoImage(display_img)                
                t1_label = Label(t1_labelframe, image=t1_display)
                t1_label.image = t1_display
                t1_label.place(x=0,y=1)
                root.update()

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"
                    
                    sheet[pos] = t1_result[i]
                
                global path3
                workbook.save(path3+"/T1.xlsx")

            if(receive_data=='C2'):
                print("Data received:", receive_data)
                t2wait_label.place_forget()
                t2_labelframe['bg'] = atk.DEFAULT_COLOR
                t2_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t2_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t2_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Processing!', font=("Courier",9,'bold'))
                tprocess_label.place(x=59,y=112)

                camera_capture(path1 + "/T2.jpg")
                
                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                t2_result, t2_image = process_image(path1 + "/T2.jpg", start_point, end_point)
                #t2_result, t2_image= process_image(path1 + "/T2.jpg")
                
                output = path2 + "/T2.jpg"
                cv2.imwrite(output, t2_image)
                t2_analysis = Image.open(output)
                t2_crop = t2_analysis.crop((start_point[0]-7, start_point[1]-7, end_point[0]+7, end_point[1]+7))

                crop_width, crop_height = t2_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t2_crop.resize((width,height))
                t2_display = ImageTk.PhotoImage(display_img)
                
                #t2_display = ImageTk.PhotoImage(t2_crop)
                t2_label = Label(t2_labelframe, image=t2_display)
                t2_label.image = t2_display
                t2_label.place(x=0,y=1)
                root.update()

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"
                    
                    sheet[pos] = t2_result[i]
                
                workbook.save(path3+"/T2.xlsx")
    
            if(receive_data=='C3'):
                print("Data received:", receive_data)
                t3wait_label.place_forget()
                t3_labelframe['bg'] = atk.DEFAULT_COLOR
                t3_labelframe['fg'] = 'lawn green'
                t_progressbar = atk.RadialProgressbar(t3_labelframe, fg='cyan')
                t_progressbar.place(x=47,y=70)
                t_progressbar.start()
                tprocess_label = Label(t3_labelframe, bg=atk.DEFAULT_COLOR, fg='white smoke', text='Processing!', font=("Courier",9,'bold'))
                tprocess_label.place(x=59,y=112)

                camera_capture(path1 + "/T3.jpg")
                
                send_data = 'C'
                ser.write(send_data.encode())
                print('Capture done!')
                t3_result, t3_image = process_image(path1 + "/T3.jpg", start_point, end_point)
                #t3result, t3_image = process_image(path1 + "/T3.jpg")
                
                output = path2 + "/T3.jpg"
                cv2.imwrite(output, t3_image)
                t3_analysis = Image.open(output)
                t3_crop = t3_analysis.crop((start_point[0]-7, start_point[1]-7, end_point[0]+7, end_point[1]+7))
                
                crop_width, crop_height = t3_crop.size
                scale_percent = 75
                width = int(crop_width * scale_percent / 100)
                height = int(crop_height * scale_percent / 100)
                display_img = t3_crop.resize((width,height))
                t3_display = ImageTk.PhotoImage(display_img)
                
                t3_label = Label(t3_labelframe, image=t3_display)
                t3_label.image = t3_display
                t3_label.place(x=0,y=1)
                wait = 0
                root.update()
                pause_button.place_forget()
                stop_button.place_forget()
                temp_label.place_forget()
                autoprocess_label.place_forget()

                sheet["A2"] = "A"
                sheet["A3"] = "B"
                sheet["A4"] = "C"
                sheet["A5"] = "D"
                sheet["A6"] = "E"
                sheet["A7"] = "F"
                sheet["A8"] = "G"
                sheet["A9"] = "H"
                sheet["B1"] = "1"
                sheet["C1"] = "2"
                sheet["D1"] = "3"
                sheet["E1"] = "4"
                sheet["F1"] = "5"
                sheet["G1"] = "6"
                for i in range(0,48):
                    if(i<6):
                        pos = str(chr(65+i+1)) + "2"
                    if(i>=6 and i<12):
                        pos = str(chr(65+i-5)) + "3"
                    if(i>=12 and i<18):
                        pos = str(chr(65+i-11)) + "4"
                    if(i>=18 and i<24):
                        pos = str(chr(65+i-17)) + "5"
                    if(i>=24 and i<30):
                        pos = str(chr(65+i-23)) + "6"
                    if(i>=30 and i<36):
                        pos = str(chr(65+i-29)) + "7"
                    if(i>=36 and i<42):
                        pos = str(chr(65+i-35)) + "8"
                    if(i>=42):
                        pos = str(chr(65+i-41)) + "9"
                    
                    sheet[pos] = t3_result[i]
                
                workbook.save(path3+"/T3.xlsx")
                
                def thr():
                    th2 = Thread(target = viewresult_click)
                    th2.start()
                def viewresult_click():
                    viewresult_button.place_forget()
                    t1_labelframe.place_forget()
                    t2_labelframe.place_forget()
                    t3_labelframe.place_forget()
                    t_labelframe.place_forget()
                    analysis_label['text']="ANALYSIS RESULTS"
                    
                    annotate_labelframe = LabelFrame(analysis_labelframe, bg='white', width=380, height=305)
                    annotate_labelframe.place(x=360,y=76)
                    root.update_idletasks()
                    
                    negative_label = Label(annotate_labelframe, bg='lawn green', width=4, height=2)
                    negative_label.place(x=75,y=32)
                    negativetext_label = Label(annotate_labelframe, bg='white', text='  (N)           NEGATIVE', height=2)
                    negativetext_label.place(x=145,y=32)
                    positive_label = Label(annotate_labelframe, bg='red', width=4, height=2)
                    positive_label.place(x=75,y=82)
                    positivetext_label = Label(annotate_labelframe, bg='white', text='  (P)           POSITIVE', height=2)
                    positivetext_label.place(x=145,y=82)
                    redue_label = Label(annotate_labelframe, bg='cyan', width=4, height=2)
                    redue_label.place(x=75,y=132)
                    reduetext_label = Label(annotate_labelframe, bg='white', text='  (R)           UNRESOLVED', height=2)
                    reduetext_label.place(x=145,y=132)
                    none_label = Label(annotate_labelframe, bg='white smoke', width=4, height=2)
                    none_label.place(x=75,y=182)
                    nonetext_label = Label(annotate_labelframe, bg='white', text='(N/A)         NOT AVAILABLE', height=2)
                    nonetext_label.place(x=145,y=182)
                    error_label = Label(annotate_labelframe, bg='yellow', width=4, height=2)
                    error_label.place(x=75,y=232)
                    errortext_label = Label(annotate_labelframe, bg='white', text='  (E)           ERROR', height=2)
                    errortext_label.place(x=145,y=232)
                    root.update_idletasks()
                    
                    def finish_click():
                        msgbox = messagebox.askquestion('Finish','Do you want to return to the main menu?', icon = 'question')
                        if(msgbox=='yes'):
                            global foldername
                            global covid19clicked
                            global tbclicked
                            foldername = ""
                            covid19clicked = 0
                            tb_clicked = 0
                            analysis_labelframe.place_forget()
                            mainscreen()                         
                    finish_button = Button(analysis_labelframe, bg="dark orange", text="FINISH", height=3, width=15, borderwidth=0, command=finish_click)
                    finish_button.place(x=480,y=396)
                    root.update_idletasks()
                    
                    result_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 307)
                    result_labelframe.place(x=104,y=120)
                    row_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=600,height = 50)
                    row_labelframe.place(x=104,y=76)
                    column_labelframe = LabelFrame(analysis_labelframe, bg='ghost white', width=50,height = 307)
                    column_labelframe.place(x=62,y=120)
                    root.update_idletasks()
                    
                    row_label = [0,0,0,0,0,0]
                    for i in range (0,6):
                        row_text = i+1
                        row_label[i] = Label(row_labelframe, text=row_text, bg='grey94', width=4, height=2)
                        row_label[i].grid(row=0,column=i,padx=2,pady=2)
                                                        
                    column_label = [0,0,0,0,0,0,0,0]
                    for i in range (0,8):
                        if(i==0):
                            column_text = 'A'
                        if(i==1):
                            column_text = 'B'
                        if(i==2):
                            column_text = 'C'
                        if(i==3):
                            column_text = 'D'
                        if(i==4):
                            column_text = 'E'
                        if(i==5):
                            column_text = 'F'
                        if(i==6):
                            column_text = 'G'
                        if(i==7):
                            column_text = 'H'
                        column_label[i] = Label(column_labelframe, text=column_text, bg='grey94', width=4, height=2)
                        column_label[i].grid(row=i,column=0,padx=2,pady=2)

                    label = list(range(48))
                    def result_table(range_a,range_b, row_value):
                        j=-1
                        global pos_result
                        for i in range(range_a, range_b):
                            j+=1
                            if(pos_result[i]<=12):
                                label[i] = Label(result_labelframe, bg='white smoke', text='N/A', width=4, height=2)
                                label[i].grid(row=row_value,column=j,padx=2,pady=2)
                            else:
                                if(t1_result[i]<=thr1_set):
                                    label[i] = Label(result_labelframe, bg='yellow', text='E', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>thr1_set and t2_result[i]<=thr2_set):
                                    label[i] = Label(result_labelframe, bg='yellow', text='E', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2) 
                                if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]<=thr3l_set):
                                    label[i] = Label(result_labelframe, bg='lawn green', text='N', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3l_set and t3_result[i]<=thr3h_set):
                                    label[i] = Label(result_labelframe, bg='cyan', text='R', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3h_set):
                                    label[i] = Label(result_labelframe, bg='red', text='P', width=4, height=2)
                                    label[i].grid(row=row_value,column=j,padx=2,pady=2)
                                                    
                    result_table(0,6,0)
                    result_table(6,12,1)
                    result_table(12,18,2)
                    result_table(18,24,3)
                    result_table(24,30,4)
                    result_table(30,36,5)
                    result_table(36,42,6)
                    result_table(42,48,7)
                    root.update_idletasks()
                    
                    def detail_click():
                        if(detail_button['bg']=='lawn green'):
                            detail_button['bg']='grey94'
                            for i in range (0,48):
                                if(pos_result[i]<=12):
                                    label[i]['text'] = 'N/A' 
                                
                                else:
                                    if(t1_result[i]<=thr1_set):
                                        label[i]['text'] = 'E' 
                                    if(t1_result[i]>thr1_set and t2_result[i]<=thr2_set):
                                        label[i]['text'] = 'E'  
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]<=thr3l):
                                        label[i]['text'] = 'N' 
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3l_set and t3_result[i]<=thr3h_set):
                                        label[i]['text'] = 'R' 
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3h_set):
                                        label[i]['text'] = 'P' 
                        else:
                            detail_button['bg']='lawn green'
                            for i in range (0,48):
                                if(pos_result[i]<=12):
                                    label[i]['text'] = str('%d'%t3_result[i]) 
                                    
                                else:
                                    if(t1_result[i]<=thr1_set):
                                        label[i]['text'] = str('%d'%t3_result[i])
                                    if(t1_result[i]>thr1_set and t2_result[i]<=thr2_set):
                                        label[i]['text'] = str('%d'%t3_result[i]) 
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]<=thr3l):
                                        label[i]['text'] = str('%d'%t3_result[i])
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3l_set and t3_result[i]<=thr3h_set):
                                        label[i]['text'] = str('%d'%t3_result[i]) 
                                    if(t1_result[i]>thr1_set and t2_result[i]>thr2_set and t3_result[i]>thr3h_set):
                                        label[i]['text'] = str('%d'%t3_result[i])
                                        
                    detail_button = Button(analysis_labelframe, activebackground="white", bg="grey94", text="DETAIL", height=3, width=10, borderwidth=0, command=detail_click)
                    detail_button.place(x=360,y=396)
                    root.update_idletasks()
                    subprocess.call(["scrot",path3+"/result.jpg"])
#                     screenshot_img = Image.open(path3+"/result.jpg")
#                     screenshot_crop = screenshot_img.crop((60,74,352,475))
#                     screenshot_crop = screenshot_crop.save(path3+"/result.jpg")
                viewresult_button = Button(analysis_labelframe, bg="dodger blue", text="VIEW RESULT", height=3, width=15, borderwidth=0, command=thr)
                viewresult_button.place(x=327,y=394)

########################################################## Khoi tao excel ###############################################################
workbook = Workbook()
sheet = workbook.active

########################################################### Serial init ################################################################ 
# ser = serial.Serial('/dev/ttyACM0', 9600)
ser = serial.Serial(
    port = '/dev/serial0',
    baudrate = 115200,
    parity = serial.PARITY_NONE,
    stopbits = serial.STOPBITS_ONE,
    bytesize = serial.EIGHTBITS,
    timeout = 1
)

############################################################## Loop ####################################################################
while True:
    mainscreen()
    root.mainloop()
